from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import os
import zipfile
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from . import models, schemas, crud
from .database import engine, get_db

load_dotenv()

# Create tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Data Manager API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Helper function to build schema response
def build_schema(view_type: str) -> schemas.TableSchema:
    """Build schema based on view type (dashboard or users)"""
    if view_type == "dashboard":
        return schemas.TableSchema(
            columns=[
                schemas.ColumnSchema(name='id', label='ID', type='text', editable=False, visible=False),
                schemas.ColumnSchema(name='name', label='Full Name', type='text', editable=True, visible=True),
                schemas.ColumnSchema(name='status', label='Status', type='select', editable=True, visible=True, options=['Active', 'Inactive', 'Pending']),
                schemas.ColumnSchema(name='department', label='Department', type='select', editable=True, visible=True, options=['Sales', 'Marketing', 'IT', 'HR', 'Finance', 'Operations', 'Legal']),
                schemas.ColumnSchema(name='salary', label='Salary', type='text', editable=True, visible=True),
                schemas.ColumnSchema(name='hire_date', label='Hire Date', type='text', editable=True, visible=False),
                schemas.ColumnSchema(name='location', label='Office Location', type='select', editable=True, visible=True, options=['New York', 'San Francisco', 'London', 'Tokyo', 'Remote']),
                schemas.ColumnSchema(name='manager', label='Manager', type='text', editable=True, visible=False),
                schemas.ColumnSchema(name='phone', label='Phone', type='text', editable=True, visible=False),
                schemas.ColumnSchema(name='employee_type', label='Employment Type', type='select', editable=True, visible=True, options=['Full-Time', 'Part-Time', 'Contract', 'Intern'])
            ]
        )
    else:  # users
        return schemas.TableSchema(
            columns=[
                schemas.ColumnSchema(name='id', label='ID', type='text', editable=False, visible=False),
                schemas.ColumnSchema(name='username', label='Username', type='text', editable=True, visible=True),
                schemas.ColumnSchema(name='email', label='Email Address', type='text', editable=True, visible=True),
                schemas.ColumnSchema(name='role', label='Role', type='select', editable=True, visible=True, options=['Admin', 'Manager', 'User', 'Guest']),
                schemas.ColumnSchema(name='status', label='Status', type='select', editable=True, visible=True, options=['Active', 'Inactive', 'Pending']),
                schemas.ColumnSchema(name='active', label='Active', type='select', editable=True, visible=True, options=['Yes', 'No']),
                schemas.ColumnSchema(name='last_login', label='Last Login', type='text', editable=True, visible=False),
                schemas.ColumnSchema(name='account_type', label='Account Type', type='select', editable=True, visible=True, options=['Premium', 'Standard', 'Trial', 'Enterprise']),
                schemas.ColumnSchema(name='created_date', label='Created Date', type='text', editable=True, visible=False),
                schemas.ColumnSchema(name='department', label='Department', type='select', editable=True, visible=True, options=['Engineering', 'Sales', 'Marketing', 'Support', 'Admin']),
                schemas.ColumnSchema(name='country', label='Country', type='select', editable=True, visible=False, options=['USA', 'UK', 'Canada', 'Germany', 'Japan', 'Australia'])
            ]
        )


def record_to_dict(record: models.Record, view_type: str) -> dict:
    """Convert record to dict based on view type"""
    base_dict = {
        'id': str(record.id),
        'name': record.name,
        'status': record.status,
        'department': record.department,
    }

    if view_type == "dashboard":
        base_dict.update({
            'salary': record.salary,
            'hireDate': record.hire_date,
            'location': record.location,
            'manager': record.manager,
            'phone': record.phone,
            'employeeType': record.employee_type,
        })
    else:  # users
        base_dict.update({
            'username': record.username,
            'email': record.email,
            'role': record.role,
            'active': record.active,
            'lastLogin': record.last_login,
            'accountType': record.account_type,
            'createdDate': record.created_date,
            'country': record.country,
        })

    return base_dict


# API Endpoints

@app.get("/")
def read_root():
    return {"message": "Data Manager API", "version": "1.0.0"}


@app.get("/api/dashboard")
def get_dashboard(
    departments: Optional[str] = None,
    statuses: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get dashboard data with optional filtering"""
    dept_list = departments.split(',') if departments else None
    status_list = statuses.split(',') if statuses else None

    records = crud.get_records(db, departments=dept_list, statuses=status_list)
    schema = build_schema("dashboard")
    data = [record_to_dict(r, "dashboard") for r in records]

    return {
        "schema": schema.model_dump(),
        "data": data
    }


@app.get("/api/users")
def get_users(
    departments: Optional[str] = None,
    statuses: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get users data with optional filtering"""
    dept_list = departments.split(',') if departments else None
    status_list = statuses.split(',') if statuses else None

    records = crud.get_records(db, departments=dept_list, statuses=status_list)
    schema = build_schema("users")
    data = [record_to_dict(r, "users") for r in records]

    return {
        "schema": schema.model_dump(),
        "data": data
    }


@app.get("/api/departments", response_model=schemas.DepartmentsResponse)
def get_departments(db: Session = Depends(get_db)):
    """Get available departments"""
    departments = crud.get_departments(db)
    return {"departments": departments}


@app.get("/api/statuses", response_model=schemas.StatusesResponse)
def get_statuses(db: Session = Depends(get_db)):
    """Get available statuses"""
    statuses = crud.get_statuses(db)
    return {"statuses": statuses}


@app.post("/api/dashboard/save", response_model=schemas.SaveResponse)
def save_dashboard(
    request: schemas.SaveRequest,
    db: Session = Depends(get_db)
):
    """Save dashboard data with transaction tracking"""
    try:
        transaction_id, records_updated = crud.update_records_with_transaction(
            db, request.data
        )
        return {
            "success": True,
            "transaction_id": transaction_id,
            "records_updated": records_updated
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/users/save", response_model=schemas.SaveResponse)
def save_users(
    request: schemas.SaveRequest,
    db: Session = Depends(get_db)
):
    """Save users data with transaction tracking"""
    try:
        transaction_id, records_updated = crud.update_records_with_transaction(
            db, request.data
        )
        return {
            "success": True,
            "transaction_id": transaction_id,
            "records_updated": records_updated
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import os
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from . import models, schemas, crud
from .database import engine, get_db
def filter_columns_by_option(data: List[dict], option: str, table_key: str) -> tuple[List[dict], List[str]]:
    """
    Filter columns based on the export option selected
    Returns: (filtered_data, column_headers)
    """
    # Define column sets for each option and table type
    # NOTE: Use camelCase names to match the frontend data format
    column_sets = {
        'users': {
            'all': ['id', 'username', 'email', 'role', 'status', 'active', 'lastLogin',
                   'accountType', 'createdDate', 'department', 'country'],
            'filtered': ['username', 'email', 'role', 'status', 'department'],  # Common fields
            'selected': ['username', 'email', 'status', 'active']  # Minimal fields
        },
        'dashboard': {
            'all': ['id', 'name', 'status', 'department', 'salary', 'hireDate',
                   'location', 'manager', 'phone', 'employeeType'],
            'filtered': ['name', 'department', 'status', 'salary', 'location'],  # Common fields
            'selected': ['name', 'department', 'status']  # Minimal fields
        }
    }

    # Get the columns for this table and option
    columns = column_sets.get(table_key, {}).get(option, [])

    # If no columns defined, return all
    if not columns:
        if data:
            columns = list(data[0].keys())

    # Filter each row to only include selected columns
    filtered_data = []
    for row in data:
        filtered_row = {col: row.get(col, '') for col in columns}
        filtered_data.append(filtered_row)

    return filtered_data, columns


def generate_excel(data: List[dict], columns: List[str], table_key: str) -> BytesIO:
    """
    Generate an Excel file from the provided data
    """
    wb = Workbook()
    ws = wb.active
    ws.title = table_key.capitalize()

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        # Convert camelCase or snake_case to Title Case
        # e.g., "lastLogin" -> "Last Login", "hire_date" -> "Hire Date"
        import re
        # Insert space before uppercase letters (camelCase) then title case
        header_text = re.sub(r'([a-z])([A-Z])', r'\1 \2', column)
        header_text = header_text.replace('_', ' ').title()
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, column in enumerate(columns, start=1):
            value = row_data.get(column, '')
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_zip(files: dict[str, BytesIO]) -> BytesIO:
    """
    Generate a ZIP file containing multiple files
    files: dict mapping filename -> BytesIO content
    """
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, content in files.items():
            content.seek(0)
            zip_file.writestr(filename, content.read())
    zip_buffer.seek(0)
    return zip_buffer


# Dashboard Version API Endpoints

@app.get("/api/versions/{view_type}")
def get_versions(
    view_type: str,
    db: Session = Depends(get_db)
):
    """Get all saved versions for a view type (dashboard or users)"""
    versions = crud.get_dashboard_versions(db, view_type)
    # Convert columns from JSONB to ColumnConfig objects
    result = []
    for v in versions:
        result.append({
            "id": str(v.id),
            "name": v.name,
            "view_type": v.view_type,
            "columns": v.columns,  # Already a list of dicts from JSONB
            "created_at": v.created_at.isoformat(),
            "updated_at": v.updated_at.isoformat()
        })
    return {"versions": result}


@app.post("/api/versions")
def create_version(
    request: schemas.DashboardVersionCreate,
    db: Session = Depends(get_db)
):
    """Create a new saved version"""
    try:
        version = crud.create_dashboard_version(db, request)
        return {
            "id": str(version.id),
            "name": version.name,
            "view_type": version.view_type,
            "columns": version.columns,
            "created_at": version.created_at.isoformat(),
            "updated_at": version.updated_at.isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/versions/{version_id}")
def update_version(
    version_id: str,
    request: schemas.DashboardVersionUpdate,
    db: Session = Depends(get_db)
):
    """Update an existing saved version"""
    try:
        from uuid import UUID
        version = crud.update_dashboard_version(db, UUID(version_id), request)
        if not version:
            raise HTTPException(status_code=404, detail="Version not found")
        return {
            "id": str(version.id),
            "name": version.name,
            "view_type": version.view_type,
            "columns": version.columns,
            "created_at": version.created_at.isoformat(),
            "updated_at": version.updated_at.isoformat()
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid version ID")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/versions/{version_id}")
def delete_version(
    version_id: str,
    db: Session = Depends(get_db)
):
    """Delete a saved version"""
    try:
        from uuid import UUID
        success = crud.delete_dashboard_version(db, UUID(version_id))
        if not success:
            raise HTTPException(status_code=404, detail="Version not found")
        return {"success": True}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid version ID")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export/{table_key}")
def export_data(
    table_key: str,
    request: schemas.ExportRequest
):
    """
    Export data to Excel file with column filtering based on option
    """
    try:
        print(f"[EXPORT] Table: {table_key}, Option: {request.option}")
        print(f"[EXPORT] Received {len(request.data)} rows")
        if request.data:
            print(f"[EXPORT] First row keys: {list(request.data[0].keys())}")

        # Filter columns based on the selected option
        filtered_data, columns = filter_columns_by_option(
            request.data,
            request.option,
            table_key
        )

        print(f"[EXPORT] After column filter: {len(filtered_data)} rows, {len(columns)} columns")
        print(f"[EXPORT] Columns: {columns}")

        # Generate Excel file
        excel_file = generate_excel(filtered_data, columns, table_key)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{table_key}_export_{request.option}_{timestamp}.xlsx"

        # Return Excel file directly
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={excel_filename}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("API_PORT", 8000))
    host = os.getenv("API_HOST", "0.0.0.0")
    uvicorn.run(app, host=host, port=port)
